import time
import openpyxl
from playwright.sync_api import Playwright, sync_playwright
from PageObjects.tutor_page import Tutorpage

class Tutoruserlogin:

    def __init__(self, page):
        self.page = page
        self.wb = openpyxl.load_workbook("C://Users//abhijit.kv//Desktop//Abhijit//FevTutor_Automation//Utilities//Upload_file//userexcel1.xlsx")
        self.sheet = self.wb['TutorLogin']
        self.tutorusernames = []
        self.tutorpasswords = []
        for i in range(10):
            username = self.sheet.cell(row=i+2, column=1).value
            password = str(self.sheet.cell(row=i+2, column=2).value)
            self.tutorusernames.append(username)
            self.tutorpasswords.append(password)
        self.login_button = page.get_by_role("button", name="Login")
        self.row = 2
   
    def read_data_from_Excelfile_tutorlogin(self):
        while self.row <= 11:
            username = self.sheet.cell(row=self.row, column=1).value
            password = str(self.sheet.cell(row=self.row, column=2).value)
            self.page.fill('input[name="USER_NAME"]', username)
            self.page.fill('input[id="Password"]', password)
            self.login_button.click(timeout=0)
            self.row += 1
            return Tutorpage(self.page)