import time
import openpyxl
from PageObjects.home_page import homepage
from playwright.sync_api import Playwright, sync_playwright

class readExcelDataPage:

    def __init__(self, page):
        self.page = page

        self.wb = openpyxl.load_workbook("C://Users//abhijit.kv//Desktop//Abhijit//FevTutor_Automation//Utilities//Upload_file//userexcel1.xlsx")
        self.sheet = self.wb['StudentLogin']
        self.usernames = []
        self.passwords = []
        for i in range(140):
            username = self.sheet.cell(row=i+2, column=1).value
            password = str(self.sheet.cell(row=i+2, column=2).value)
            self.usernames.append(username)
            self.passwords.append(password)
        self.login_button = page.get_by_role("button", name="Login")
        self.row = 2

    def read_data_from_Excelfile_studentlogin(self):
        while self.row <= 140:
            username = self.sheet.cell(row=self.row, column=1).value
            password = str(self.sheet.cell(row=self.row, column=2).value)
            self.page.fill('input[name="USER_NAME"]', username)
            self.page.fill('input[id="Password"]', password)
            self.login_button.click(timeout=0)
            self.row += 1
            return homepage(self.page)
        

# import pandas as pd

 

# # Load the Excel file into a DataFrame
# file_path = "C://Users//abhijit.kv//Desktop//Abhijit//FevTutor_Automation//Utilities//Upload_file//userexcel1.xlsx"
# df = pd.read_excel(file_path) 

# # Input from the user
# user_id = input("username")
# password = input("password")

 

# # Check if the provided credentials are in the DataFrame
# if (df['User ID'] == user_id).any() and (df['Password'] == password).any():
#     print("Login successful!")
# else:
#     print("Login failed. Please check your user ID and password.")