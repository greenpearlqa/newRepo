import openpyxl
from datetime import datetime, timedelta

from Utilities.paths import get_bulk_plan_upload_file

class BulkplanPage:
    def __init__(self,page):
        self.page = page
         
        self.excel_file_path = get_bulk_plan_upload_file()
        self.workbook = openpyxl.load_workbook(self.excel_file_path)
        self.sheet = self.workbook.active

    def update_data(self):
        row_index=2
        current_date = datetime.now().strftime('%d-%b-%Y')
        # Update data for a specific row
        #current_date = datetime.now()
        self.sheet.cell(row=row_index, column=1, value='FevQAStu1021')
        self.sheet.cell(row=row_index, column=2, value='18781')
        self.sheet.cell(row=row_index, column=3, value='')
        self.sheet.cell(row=row_index, column=4, value='')
        self.sheet.cell(row=row_index, column=5, value='(GMT-05:00) Eastern Time (US & Canada)')
        self.sheet.cell(row=row_index, column=6, value='')
        self.sheet.cell(row=row_index, column=7, value='')

        self.sheet.cell(row=row_index, column=8, value=current_date)
        self.sheet.cell(row=row_index, column=9, value=(datetime.now() + timedelta(days=5)).strftime('%d-%b-%Y'))
        self.sheet.cell(row=row_index, column=10, value='No')
         # Calculate the day dynamically
        day_of_week = datetime.now().strftime('%A')  # Get the full day name
        self.sheet.cell(row=row_index, column=11, value=day_of_week)
        self.sheet.cell(row=row_index, column=12, value='2')
        self.sheet.cell(row=row_index, column=13, value='')
        self.sheet.cell(row=row_index, column=14, value='16:00')
        self.sheet.cell(row=row_index, column=15, value='60')

        # Save the workbook
        self.workbook.save(self.excel_file_path)
        self.workbook.close()
        self.page.wait_for_timeout(2000)









